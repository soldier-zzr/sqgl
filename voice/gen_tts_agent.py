# -*- coding: utf-8 -*-
"""
话术生成智能体（DeepSeek 版）

功能：
  1. 读取「话术原料」Excel（姓名列 + 赛道列）
  2. 读取「好友数据」Excel（昵称A列 + 备注B列）
  3. 按姓名=昵称匹配，用 DeepSeek AI 为每人生成个性化话术
  4. 输出 tts_input.txt（格式：备注::话术），直接粘贴进 TTS 工具

用法：
  pip install openai openpyxl
  python gen_tts_agent.py
"""

import sys
import time
from pathlib import Path

# ======================== 配置区 ========================

_BASE = Path(__file__).resolve().parent

# 话术原料 Excel（含「姓名/昵称」列 + 「赛道」列 + 可选「备注」列）
MATERIAL_EXCEL = _BASE / "话术原料.xlsx"

# 输出文件
OUTPUT_TXT = _BASE / "tts_input.txt"

# DeepSeek API 配置（优先从 config.py 读取，不存在则使用下方默认值）
DEEPSEEK_BASE_URL = "https://api.deepseek.com"
MODEL = "deepseek-chat"
try:
    from config import DEEPSEEK_API_KEY  # type: ignore
except ImportError:
    DEEPSEEK_API_KEY = ""  # 请在 config.py 中设置

# 每次 API 调用之间的间隔（秒），避免触发限流
API_DELAY = 0.3

# ========================================================

NAME_ALIASES   = {"姓名", "昵称", "名字", "name"}
TRACK_ALIASES  = {"赛道", "行业", "领域", "niche", "track", "方向"}
REMARK_ALIASES = {"备注", "remark", "备注名", "mp3名"}

# 提示词从同目录 prompt.txt 读取，不存在则使用内置默认值
_PROMPT_FILE = Path(__file__).resolve().parent / "prompt.txt"
if _PROMPT_FILE.exists():
    SYSTEM_PROMPT = _PROMPT_FILE.read_text(encoding="utf-8").strip()
else:
    SYSTEM_PROMPT = """你是一个专业的个人IP营销文案专家。
你的任务是为每位学员生成一段个性化的直播预约语音话术。

规则：
1. 话术要自然口语化，像真人录音一样
2. 有赛道时使用个性化模板，无赛道时使用通用模板
3. 话术长度控制在100字以内
4. 直接输出话术文本，不要加任何前缀、序号或解释

个性化模板（有赛道时参考）：
[姓名]老师，你好啊，这是一条专门为你录的语音。我看你是做[赛道]赛道的，我们有很多博主都在做这个赛道。今天晚上我会讲21天IP起盘地图，用AI就能快速把IP做起来，到时候会分享一些跟你相关的案例，你可以关注一下。晚上8点见。

通用模板（无赛道时参考）：
[姓名]老师，我提示一下啊，今天晚上8点，咱第二天分享21天快速打造一个IP的地图，是我们26年核心打法，我建议你听一下，会对你做的赛道有帮助，有很多跟你相关的一些案例啊，晚上8点见一会儿见。"""


def find_col(headers: list, aliases: set) -> int | None:
    for i, h in enumerate(headers):
        if str(h or "").strip().lower() in {a.lower() for a in aliases}:
            return i
    return None


def load_contacts(path: Path) -> dict[str, str]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nickname = str(row[0]).strip() if row[0] else ""
        remark   = str(row[1]).strip() if row[1] else ""
        if nickname and remark:
            result[nickname] = remark
    return result


def load_material(path: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    name_col  = find_col(headers, NAME_ALIASES)
    track_col = find_col(headers, TRACK_ALIASES)

    if name_col is None:
        print(f"[ERR] 话术原料表未找到姓名列，当前列头：{headers}")
        print("      请在配置区 NAME_ALIASES 中添加对应列名")
        sys.exit(1)

    print(f"  姓名列：{headers[name_col]}（第{name_col+1}列）")
    if track_col is not None:
        print(f"  赛道列：{headers[track_col]}（第{track_col+1}列）")
    else:
        print("  未找到赛道列，将全部使用通用话术模板")

    remark_col = find_col(headers, REMARK_ALIASES)
    if remark_col is not None:
        print(f"  备注列：{headers[remark_col]}（第{remark_col+1}列）")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_col] or "").strip()
        if not name or name.lower() == "none":
            continue
        track = ""
        if track_col is not None and row[track_col]:
            t = str(row[track_col]).strip()
            if t and t.lower() not in ("none", "无", "未知", "-", ""):
                track = t
        remark = ""
        if remark_col is not None and row[remark_col]:
            remark = str(row[remark_col]).strip()
        rows.append({"name": name, "track": track, "remark": remark})
    return rows


def generate_script(client, name: str, track: str) -> str:
    if track:
        user_msg = f"姓名：{name}\n赛道：{track}\n\n请使用个性化模板生成话术。"
    else:
        user_msg = f"姓名：{name}\n赛道：无\n\n请使用通用模板生成话术。"

    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_msg},
        ],
        max_tokens=300,
        temperature=0.7,
    )
    return resp.choices[0].message.content.strip()


def main():
    try:
        from openai import OpenAI
        import openpyxl  # noqa
    except ImportError as e:
        print(f"[ERR] 缺少依赖：{e}")
        print("请运行：pip install openai openpyxl")
        sys.exit(1)

    for f in [CONTACTS_EXCEL, MATERIAL_EXCEL]:
        if not f.exists():
            print(f"[ERR] 文件不存在: {f}")
            sys.exit(1)

    client = OpenAI(api_key=DEEPSEEK_API_KEY, base_url=DEEPSEEK_BASE_URL)

    print(f"读取好友数据: {CONTACTS_EXCEL.name}")
    contacts = load_contacts(CONTACTS_EXCEL)
    print(f"  共 {len(contacts)} 条昵称→备注映射")

    print(f"\n读取话术原料: {MATERIAL_EXCEL.name}")
    material = load_material(MATERIAL_EXCEL)
    print(f"  共 {len(material)} 条待生成")

    lines = []
    skipped = []
    failed = []

    print(f"\n开始生成话术（模型：{MODEL}）...\n")

    for i, item in enumerate(material):
        name  = item["name"]
        track = item["track"]

        remark = contacts.get(name)
        if remark is None:
            remark = next((v for k, v in contacts.items() if k.lower() == name.lower()), None)
        if remark is None:
            print(f"  [{i+1}/{len(material)}] {name} - 跳过（好友表中未找到）")
            skipped.append(name)
            continue

        try:
            script = generate_script(client, name, track)
            script = script.replace("\n", " ").replace("\r", "").strip()
            lines.append(f"{remark}::{script}")
            tag = f"({track})" if track else "(通用)"
            print(f"  [{i+1}/{len(material)}] {name} {tag} - OK")
            if API_DELAY > 0 and i < len(material) - 1:
                time.sleep(API_DELAY)
        except Exception as e:
            print(f"  [{i+1}/{len(material)}] {name} - 失败: {e}")
            failed.append(name)

    OUTPUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    print(f"\n{'='*50}")
    print(f"生成完成 -> {OUTPUT_TXT}")
    print(f"  成功：{len(lines)} 条")
    if skipped:
        print(f"  跳过（好友表未找到）：{skipped}")
    if failed:
        print(f"  失败（API错误）：{failed}")
    print(f"\n将 {OUTPUT_TXT.name} 内容粘贴进 TTS 工具输入框，每行一条任务。")


if __name__ == "__main__":
    main()
