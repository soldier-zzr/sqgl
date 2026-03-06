# -*- coding: utf-8 -*-
"""
TTS 输入文本生成器

用法：
  python gen_tts_input.py

功能：
  1. 读取「好友数据」Excel（昵称A列 + 备注B列）
  2. 读取「豆包话术」Excel（姓名列 + 生成话术列）
  3. 按姓名=昵称匹配，拼出 备注::话术 格式
  4. 输出 tts_input.txt，直接粘贴进 TTS 工具即可

配置区：修改下方三个路径即可
"""

from pathlib import Path
import sys

# ======================== 配置区 ========================

# 好友数据 Excel（A列=昵称，B列=备注）
CONTACTS_EXCEL = Path(r"C:\Users\11424\Desktop\新建文件夹\voice\好友数据_20260305T125413.xlsx")

# 豆包输出的话术 Excel（含姓名列 + 生成话术列）
DOUBAN_EXCEL = Path(r"C:\Users\11424\Desktop\新建文件夹\voice\话术.xlsx")

# 输出文件（直接粘贴进 TTS 工具）
OUTPUT_TXT = Path(r"C:\Users\11424\Desktop\新建文件夹\voice\tts_input.txt")

# ========================================================

# 豆包 Excel 中，姓名列可能叫「姓名」「昵称」「名字」等
NAME_ALIASES = {"姓名", "昵称", "名字", "name"}
# 话术列可能叫「生成话术」「话术」「文案」「内容」等
SCRIPT_ALIASES = {"生成话术", "话术", "文案", "内容", "script", "text"}


def find_col(headers: list[str], aliases: set[str]) -> int | None:
    """不区分大小写、去空格，找目标列的索引"""
    for i, h in enumerate(headers):
        if str(h).strip().lower() in {a.lower() for a in aliases}:
            return i
    return None


def load_contacts(path: Path) -> dict[str, str]:
    """返回 {昵称: 备注} 映射"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nickname = str(row[0]).strip() if row[0] else ""
        remark   = str(row[1]).strip() if row[1] else ""
        if nickname and remark:
            result[nickname] = remark
    return result


def load_scripts(path: Path) -> list[tuple[str, str]]:
    """返回 [(姓名, 话术), ...]"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 第一行是表头
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col   = find_col(headers, NAME_ALIASES)
    script_col = find_col(headers, SCRIPT_ALIASES)

    if name_col is None:
        print(f"[ERR] 未找到姓名列，当前列头：{headers}")
        print(f"      请在配置区 NAME_ALIASES 中添加对应列名")
        sys.exit(1)
    if script_col is None:
        print(f"[ERR] 未找到话术列，当前列头：{headers}")
        print(f"      请在配置区 SCRIPT_ALIASES 中添加对应列名")
        sys.exit(1)

    print(f"  姓名列：{headers[name_col]}（第{name_col+1}列）")
    print(f"  话术列：{headers[script_col]}（第{script_col+1}列）")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name   = str(row[name_col]).strip()   if row[name_col]   else ""
        script = str(row[script_col]).strip() if row[script_col] else ""
        if name and script and name.lower() != "none" and script.lower() != "none":
            rows.append((name, script))
    return rows


def main():
    try:
        import openpyxl  # noqa
    except ImportError:
        print("[ERR] 缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    # 检查文件
    for f in [CONTACTS_EXCEL, DOUBAN_EXCEL]:
        if not f.exists():
            print(f"[ERR] 文件不存在: {f}")
            sys.exit(1)

    print(f"读取好友数据: {CONTACTS_EXCEL.name}")
    contacts = load_contacts(CONTACTS_EXCEL)
    print(f"  共 {len(contacts)} 条昵称→备注映射")

    print(f"\n读取话术数据: {DOUBAN_EXCEL.name}")
    scripts = load_scripts(DOUBAN_EXCEL)
    print(f"  共 {len(scripts)} 条话术")

    # 合并
    lines = []
    skipped = []
    for name, script in scripts:
        remark = contacts.get(name)
        if remark is None:
            # 尝试忽略大小写匹配
            remark = next((v for k, v in contacts.items() if k.lower() == name.lower()), None)
        if remark:
            # 话术里换行符替换为空格，保证单行
            clean_script = script.replace("\n", " ").replace("\r", "").strip()
            lines.append(f"{remark}::{clean_script}")
        else:
            skipped.append(name)

    # 写输出
    OUTPUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    print(f"\n生成完成：{OUTPUT_TXT}")
    print(f"  成功：{len(lines)} 条")
    if skipped:
        print(f"  跳过（好友表中未找到）：{len(skipped)} 条")
        for n in skipped:
            print(f"    - {n}")

    print(f"\n请将 {OUTPUT_TXT.name} 的内容粘贴进 TTS 工具输入框，每行一条任务。")


if __name__ == "__main__":
    main()
